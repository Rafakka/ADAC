import os
from datetime import datetime
from config import CONTATOS_DIR

import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class XLSXManager:
    def __init__(self, xlsx_path=None):
        if xlsx_path is None:
            self.xlsx_path = self.encontrar_arquivo_xlsx()
        else:
            self.xlsx_path = xlsx_path

        self.arquivo_existe = os.path.exists(self.xlsx_path)

    def encontrar_arquivo_xlsx(self):
        """Encontra automaticamente o arquivo XLSX na pasta contatos/"""
        try:
            if os.path.exists(CONTATOS_DIR):
                for arquivo in os.listdir(CONTATOS_DIR):
                    if arquivo.lower().endswith('.xlsx'):
                        return os.path.join(CONTATOS_DIR, arquivo)
            return None
        except Exception as e:
            print(f"Erro ao procurar arquivo XLSX: {e}")
            return None

    def arquivo_existente(self):
        return self.arquivo_existe

    def ler_contatos(self):
        """Lê contatos do XLSX"""
        if not self.arquivo_existe:
            return []

        contatos = []
        try:
            wb = load_workbook(self.xlsx_path)
            ws = wb.active

            # Cabeçalhos (linhas de título)
            headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

            for row in ws.iter_rows(min_row=2, values_only=False):
                paciente = row[headers['PACIENTE'] - 1].value
                data_nasc = row[headers['DATA NASCIMENTO'] - 1].value
                tel_recado = row[headers['TEL. RECADO'] - 1].value
                tel_celular = row[headers['TEL. CELULAR'] - 1].value
                status = row[headers.get('status', 0) - 1].value if 'status' in headers else None

                if status not in ['SUCESSO', 'PROCESSADO']:
                    numero = tel_recado if tel_recado else tel_celular
                    contatos.append({
                        'numero': str(numero) if numero else '',
                        'nome': paciente or '',
                        'data_nascimento': str(data_nasc) if data_nasc else '',
                        'status': status or 'PENDENTE'
                    })

            return contatos
        except Exception as e:
            print(f"Erro ao ler XLSX: {e}")
            return []

    def marcar_como_processado(self, numero, status, nome="", data_nascimento=""):
        """Atualiza status no XLSX e pinta de azul/vermelho"""
        if not self.arquivo_existe:
            return False

        try:
            wb = load_workbook(self.xlsx_path)
            ws = wb.active

            headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

            for row in ws.iter_rows(min_row=2):
                tel_recado = row[headers['tel. recado'] - 1].value
                tel_celular = row[headers['tel. celular'] - 1].value
                numero_atual = str(tel_recado if tel_recado else tel_celular)

                if numero_atual == str(numero):
                    # Atualizar status e data_processamento
                    if 'status' in headers:
                        row[headers['status'] - 1].value = status
                    if 'data_processamento' in headers:
                        row[headers['data_processamento'] - 1].value = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

                    # Pintar célula do status
                    fill = None
                    if status.upper() in ["SUCESSO", "ATENDEU"]:
                        fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Azul claro
                    else:
                        fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")  # Vermelho tomate

                    if fill and 'status' in headers:
                        row[headers['status'] - 1].fill = fill

                    # Atualizar nome e data_nascimento se fornecidos
                    if nome and 'paciente' in headers:
                        row[headers['paciente'] - 1].value = nome
                    if data_nascimento and 'DATA NASCIMENTO' in headers:
                        row[headers['DATA NASCIMENTO'] - 1].value = data_nascimento

            wb.save(self.xlsx_path)
            return True
        except Exception as e:
            print(f"Erro ao atualizar XLSX: {e}")
            return False

    def get_xlsx_path(self):
        return self.xlsx_path
